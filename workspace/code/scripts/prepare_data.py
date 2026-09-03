from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[3]
RAW = ROOT / "workspace" / "data_raw"
CLEAN = ROOT / "workspace" / "data_clean"
PROFILE_DIR = ROOT / "workspace" / "data"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def clean_text(value):
    if not isinstance(value, str):
        return value
    return re.sub(r"\s+", " ", value).strip()


def merged_value_map(ws):
    values = {}
    for merged in ws.merged_cells.ranges:
        anchor = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                values[(row, col)] = anchor
    return values


def read_rows(ws, columns, numeric_key_col=1):
    merged = merged_value_map(ws)
    rows = []
    for row_no in range(2, ws.max_row + 1):
        key = merged.get((row_no, numeric_key_col), ws.cell(row_no, numeric_key_col).value)
        if key is None:
            continue
        row = []
        for col in range(1, columns + 1):
            value = merged.get((row_no, col), ws.cell(row_no, col).value)
            row.append(clean_text(value))
        rows.append(row)
    return rows


def write_csv(path: Path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(headers)
        writer.writerows(rows)


def parse_price_range(value: str):
    match = re.fullmatch(r"\s*([0-9.]+)\s*-\s*([0-9.]+)\s*", str(value))
    if not match:
        raise ValueError(f"Unrecognized price range: {value!r}")
    low, high = map(float, match.groups())
    if low > high:
        raise ValueError(f"Price range is reversed: {value!r}")
    return low, high


def duplicate_count(rows, indexes):
    keys = [tuple(row[index] for index in indexes) for row in rows]
    return len(keys) - len(set(keys))


def main():
    CLEAN.mkdir(parents=True, exist_ok=True)
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    book1_path = RAW / "附件1.xlsx"
    book2_path = RAW / "附件2.xlsx"
    book1 = openpyxl.load_workbook(book1_path, data_only=False)
    book2 = openpyxl.load_workbook(book2_path, data_only=False)

    land_raw = read_rows(book1["乡村的现有耕地"], 4)
    land = [[row[0], row[1], float(row[2]), row[3]] for row in land_raw]
    crop_raw = read_rows(book1["乡村种植的农作物"], 5)
    crops = [row for row in crop_raw if isinstance(row[0], (int, float))]
    crops = [[int(row[0]), row[1], row[2], row[3], row[4]] for row in crops]

    planting_raw = read_rows(book2["2023年的农作物种植情况"], 6)
    planting = [row for row in planting_raw if isinstance(row[1], (int, float))]
    planting = [[row[0], int(row[1]), row[2], row[3], float(row[4]), row[5]] for row in planting]

    stats_raw = read_rows(book2["2023年统计的相关数据"], 8)
    stats = [row for row in stats_raw if isinstance(row[1], (int, float))]
    stats_clean = []
    price_widths = []
    for row in stats:
        low, high = parse_price_range(row[7])
        price_widths.append(high - low)
        stats_clean.append([
            int(row[0]), int(row[1]), row[2], row[3], row[4],
            float(row[5]), float(row[6]), low, high,
        ])

    write_csv(CLEAN / "land.csv", ["plot_id", "land_type", "area_mu", "source_note"], land)
    write_csv(
        CLEAN / "crops.csv",
        ["crop_id", "crop_name", "crop_type", "source_suitable_land", "source_note"],
        crops,
    )
    write_csv(
        CLEAN / "planting_2023.csv",
        ["plot_id", "crop_id", "crop_name", "crop_type", "area_mu", "season"],
        planting,
    )
    write_csv(
        CLEAN / "economics_2023.csv",
        [
            "record_id", "crop_id", "crop_name", "land_type", "season",
            "yield_jin_per_mu", "cost_yuan_per_mu", "price_low_yuan_per_jin",
            "price_high_yuan_per_jin",
        ],
        stats_clean,
    )

    land_area = {row[0]: row[2] for row in land}
    planting_sums = defaultdict(float)
    unknown_plots = []
    unknown_crops = []
    crop_ids = {row[0] for row in crops}
    for row in planting:
        plot, crop_id, _, _, area, season = row
        if plot not in land_area:
            unknown_plots.append(plot)
        if crop_id not in crop_ids:
            unknown_crops.append(crop_id)
        planting_sums[(plot, season)] += area
    area_excess = [
        {"plot_id": plot, "season": season, "planted_area_mu": total, "plot_area_mu": land_area[plot]}
        for (plot, season), total in planting_sums.items()
        if plot in land_area and total > land_area[plot] + 1e-9
    ]

    missing_economic_combinations = []
    economic_crop_ids = {row[1] for row in stats_clean}
    for crop_id in sorted(crop_ids - economic_crop_ids):
        missing_economic_combinations.append(crop_id)
    smart_first_season_missing = sorted(
        crop_id for crop_id in range(17, 35)
        if not any(row[1] == crop_id and row[3] == "智慧大棚" and row[4] == "第一季" for row in stats_clean)
    )

    land_type_counts = Counter(row[1] for row in land)
    crop_type_counts = Counter(row[2] for row in crops)
    planting_crop_counts = Counter(row[1] for row in planting)
    stats_land_counts = Counter(row[3] for row in stats_clean)

    raw_files = []
    for path in [RAW / "C题.pdf", book1_path, book2_path]:
        raw_files.append({
            "path": path.relative_to(ROOT).as_posix(),
            "size_bytes": path.stat().st_size,
            "sha256": sha256(path),
        })

    profile = {
        "schema_version": 1,
        "raw_files": raw_files,
        "attachment_mapping": [
            {"file": "C题.pdf", "scope": ["Q1", "Q2", "Q3"], "role": "problem statement"},
            {"file": "附件1.xlsx", "scope": ["Q1", "Q2", "Q3"], "role": "land and crop master data"},
            {"file": "附件2.xlsx", "scope": ["Q1", "Q2", "Q3"], "role": "2023 planting and economics"},
            {"file": "附件3", "scope": ["Q1", "Q2"], "role": "required output templates", "status": "missing"},
        ],
        "fields": [
            {"dataset": "land", "key": ["plot_id"], "units": {"area_mu": "亩"}},
            {"dataset": "crops", "key": ["crop_id"], "units": {}},
            {"dataset": "planting_2023", "key": ["plot_id", "crop_id", "season"], "units": {"area_mu": "亩"}},
            {
                "dataset": "economics_2023",
                "key": ["crop_id", "land_type", "season"],
                "units": {
                    "yield_jin_per_mu": "斤/亩",
                    "cost_yuan_per_mu": "元/亩",
                    "price_low_yuan_per_jin": "元/斤",
                    "price_high_yuan_per_jin": "元/斤"
                }
            }
        ],
        "quality": {
            "missingness": {
                "required_fields_after_merged_cell_expansion": 0,
                "note": "Source presentation uses merged cells; inherited values were expanded in cleaned tables."
            },
            "duplicates": {
                "land_primary_key": duplicate_count(land, [0]),
                "crop_primary_key": duplicate_count(crops, [0]),
                "planting_composite_key": duplicate_count(planting, [0, 1, 5]),
                "economics_composite_key": duplicate_count(stats_clean, [1, 3, 4])
            },
            "impossible_values": {
                "nonpositive_land_area": sum(row[2] <= 0 for row in land),
                "nonpositive_planting_area": sum(row[4] <= 0 for row in planting),
                "nonpositive_yield": sum(row[5] <= 0 for row in stats_clean),
                "negative_cost": sum(row[6] < 0 for row in stats_clean),
                "nonpositive_price_bound": sum(row[7] <= 0 or row[8] <= 0 for row in stats_clean),
                "plot_season_area_excess": area_excess,
                "unknown_plot_references": sorted(set(unknown_plots)),
                "unknown_crop_references": sorted(set(unknown_crops)),
                "crop_ids_without_economic_rows": missing_economic_combinations
                ,"economic_slot_coverage_gaps": {
                    "smart_greenhouse_first_season_crop_ids": smart_first_season_missing,
                    "note": "Attachment 1 permits two smart-greenhouse vegetable seasons and 2023 planting includes first-season records, but Attachment 2 supplies only smart-greenhouse rows labeled second season."
                }
            },
            "outliers": {
                "status": "not_labeled_as_errors",
                "note": "Large yield, cost and price differences are crop-dependent and must not be globally winsorized. Method-specific screening should inspect profitability concentration."
            }
        },
        "coverage": {
            "rows": {"land": len(land), "crops": len(crops), "planting_2023": len(planting), "economics_2023": len(stats_clean)},
            "effective_sample_size": {
                "plots": len(land),
                "crops": len(crops),
                "planting_records": len(planting),
                "economic_records": len(stats_clean)
            },
            "time_range": [2023, 2023],
            "time_gaps": None,
            "time_note": "Only a 2023 cross-section is supplied; future behavior is scenario-driven, not estimable as a historical time series."
        },
        "distribution_risks": {
            "class_imbalance": {
                "land_type_counts": dict(sorted(land_type_counts.items())),
                "crop_type_counts": dict(sorted(crop_type_counts.items())),
                "2023_planting_records_per_crop_min": min(planting_crop_counts.values()),
                "2023_planting_records_per_crop_max": max(planting_crop_counts.values()),
                "economic_records_by_land_type": dict(sorted(stats_land_counts.items()))
            },
            "rare_categories": [
                "智慧大棚仅4个地块、总面积2.4亩。",
                "食用菌仅4种，且只允许普通大棚第二季。"
            ],
            "high_cardinality": ["plot_id: 54 levels", "crop_id: 41 levels"],
            "redundancy_warnings": [
                "crop_name and crop_type repeat information keyed by crop_id; joins should use crop_id and validate names."
            ],
            "concentration_metrics": {
                "price_range_width_min": min(price_widths),
                "price_range_width_max": max(price_widths),
                "note": "No outcome target exists yet; decision-output concentration must be checked in method risk probes."
            }
        },
        "per_question_readiness": {
            "Q1": {
                "status": "ready_with_warnings",
                "warnings": [
                    "Expected sales volume is not supplied as a separate field.",
                    "Sales price is a range and no representative-value rule is specified.",
                    "Smart-greenhouse first-season economics are not separately supplied.",
                    "Management thresholds and Attachment 3 templates are missing."
                ]
            },
            "Q2": {
                "status": "ready_with_warnings",
                "warnings": [
                    "Only 2023 data are observed; future distributions and risk tolerance require explicit scenario assumptions.",
                    "Smart-greenhouse first-season economics are not separately supplied.",
                    "Attachment 3 output template is missing."
                ]
            },
            "Q3": {
                "status": "ready_with_warnings",
                "warnings": [
                    "The one-year supplied data cannot identify substitution, complementarity, or temporal correlations.",
                    "Smart-greenhouse first-season economics are not separately supplied.",
                    "Simulation relationships and parameters require literature support or human-approved assumptions."
                ]
            }
        },
        "cleaned_files": [
            {"path": "workspace/data_clean/land.csv", "transformations": ["expanded merged cells", "trimmed whitespace"]},
            {"path": "workspace/data_clean/crops.csv", "transformations": ["expanded merged cells", "trimmed whitespace"]},
            {"path": "workspace/data_clean/planting_2023.csv", "transformations": ["expanded merged plot identifiers", "trimmed whitespace"]},
            {"path": "workspace/data_clean/economics_2023.csv", "transformations": ["trimmed whitespace", "parsed price range into low and high bounds"]}
        ],
        "unresolved_risks": [
            "No Attachment 3 templates.",
            "Expected sales volume proxy not confirmed.",
            "Price representative value not confirmed.",
            "Management convenience thresholds not confirmed.",
            "Smart-greenhouse first-season yield/cost rows are absent; whether second-season parameters apply to both seasons is not explicitly stated.",
            "Q2 probability and risk conventions not confirmed.",
            "Q3 relationship groups and dependence parameters are not empirically identifiable from supplied data."
        ]
    }

    (PROFILE_DIR / "data_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


if __name__ == "__main__":
    main()
